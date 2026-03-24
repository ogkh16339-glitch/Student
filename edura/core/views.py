from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages 
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Q
from django.http import JsonResponse
from datetime import date

from datetime import datetime

# Modellarni import qilish
from student_management.models import News, Comment, Student, Attendance, VideoNews
from .models import ContactMessage
from .forms import NewsForm, ContactForm
from .models import ContactMessage, Timetable
# finance/models.py ichidan Payment va StudentBalance ni import qilish
from finance.models import Payment, StudentBalance  # <-- SHU QATORNI TEKSHIRING
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib import messages
# Pastdagi qatorni tekshiring, Student va Payment mavjudligiga ishonch hosil qiling
from student_management.models import Student  # Talaba boshqa appda bo'lgani uchun
from finance.models import Payment
from django.db import transaction
from .forms import TimetableForm
# 1. DASHBOARD VA ASOSIY
from django.shortcuts import render
from .models import Timetable  # Timetable o'zining ilovasida (core)
from student_management.models import Group, Teacher  # Boshqa appdan import qilish
from .forms import TimetableForm
from datetime import datetime
from .models import Task
import openpyxl
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

# views.py
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from student_management.models import News, Bookmark 

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db import transaction, models # Yangilik: tranzaksiya va Q query uchun
from .models import Timetable
from student_management.models import Group, Teacher
from datetime import datetime


from django.db.models import Q
from datetime import datetime

@login_required(login_url='core:login')
def dashboard(request):
    if not request.user.is_superuser:
        return redirect('users:student_profile')

    # Formalarni inisializatsiya qilish
    news_form = NewsForm() 
    timetable_form = TimetableForm()

    # --- POST SO'ROVLARI ---
    if request.method == 'POST':
        if 'add_news' in request.POST:
            news_form = NewsForm(request.POST, request.FILES)
            if news_form.is_valid():
                news = news_form.save(commit=False)
                news.author = request.user
                news.save()
                messages.success(request, "Yangilik muvaffaqiyatli chop etildi!")
                return redirect('core:dashboard')
        
        elif 'add_timetable' in request.POST:
            timetable_form = TimetableForm(request.POST)
            if timetable_form.is_valid():
                timetable_form.save()
                messages.success(request, "Dars jadvali muvaffaqiyatli qo'shildi!")
                return redirect('core:dashboard')

    # --- MA'LUMOTLARNI YIG'ISH ---
    today = date.today()
    
    # 1. Davomat va Talabalar
    active_students = Student.objects.filter(enrollment_status='active')
    total_students = active_students.count()
    
    attendance_today = Attendance.objects.filter(date=today)
    present_today = attendance_today.filter(status='present').count()
    absent_today = attendance_today.filter(status='absent').count()
    late_today = attendance_today.filter(status='late').count()
    
    attendance_rate = (present_today / total_students * 100) if total_students > 0 else 0

    # 2. Xabarlar (Contact Messages)
    all_active_messages = ContactMessage.objects.filter(is_deleted=False)
    tasks = Task.objects.filter(user=request.user)[:10]
    
    # 3. Jadval (Timetable)
    day_names = {1: 'Dushanba', 2: 'Seshanba', 3: 'Chorshanba', 4: 'Payshanba', 5: 'Juma', 6: 'Shanba', 7: 'Yakshanba'}
    timetable_data = {}
    for num, name in day_names.items():
        lessons = Timetable.objects.filter(day_of_week=num).select_related('group').order_by('start_time')
        if lessons.exists():
            timetable_data[name] = lessons

    # --- CONTEXT ---
    context = {
        # Statistika kartalari
        'total_students': total_students,
        'present_today': present_today,
        'absent_today': absent_today,
        'late_today': late_today,
        'attendance_rate': round(attendance_rate, 1),
        
        # Ro'yxatlar
        'recent_students': Student.objects.all().order_by('-id')[:5],
        'latest_news': News.objects.all().order_by('-created_at')[:5],
        'timetable_data': timetable_data,
        'today_name': day_names.get(datetime.now().isoweekday()),
        
        # Xabarlar qismi (Sizning kodingizda bor edi)
        'user_messages': all_active_messages.order_by('-created_at')[:3],
        'recent_contact_messages': all_active_messages.order_by('-created_at')[:5],
        'total_msg_count': all_active_messages.count(),
        'unread_msg_count': all_active_messages.filter(is_read=False).count(),
        'today_msg_count': all_active_messages.filter(created_at__date=today).count(),
        'tasks': tasks,
        
        # Formalar
        'news_form': news_form,
        'timetable_form': timetable_form,
    }
    
    return render(request, 'main/index.html', context)




def delete_timetable(request, id):
    # Darsni bazadan qidiramiz, topilmasa 404 beradi
    lesson = get_object_or_404(Timetable, id=id)
    lesson.delete()
    return JsonResponse({'status': 'success'})

def home(request):
    return render(request, 'core/home.html')

def about(request):
    return render(request, 'core/about.html')


# 2. AVTORIZATSIYA


def admin_login(request):
    if request.user.is_authenticated:
        return redirect('core:dashboard')

    if request.method == 'POST':
        user = request.POST.get('username')
        passw = request.POST.get('password')
        admin_user = authenticate(username=user, password=passw)
        
        if admin_user is not None:
            login(request, admin_user)
            return redirect('core:dashboard')
        else:
            return render(request, 'core/login.html', {'error': "Login yoki parol noto'g'ri!"})
            
    return render(request, 'core/login.html')

def logout_view(request):
    logout(request)
    return redirect('core:login')

@login_required(login_url='core:login')
def admin_profile(request):
    return render(request, 'core/profile.html')


# 3. YANGILIKLAR VA MEDIATEKA

from student_management.models import News, YouTubeVideo
from django.http import JsonResponse # AJAX uchun kerak

def news_list(request):
    # Sizning mavjud kodingiz (o'zgarishsiz qoladi)
    query = request.GET.get('q')
    news_qs = News.objects.all().order_by('-created_at')
    
    if query:
        news_qs = news_qs.filter(Q(title__icontains=query) | Q(content__icontains=query))
    
    videos = VideoNews.objects.all().order_by('-created_at')
    paginator = Paginator(news_qs, 6) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    yt_video = YouTubeVideo.objects.filter(is_active=True).last()
    news = News.objects.all().order_by('-created_at')
    
    return render(request, 'core/news_list.html', {
        'news': news,
        'yt_video': yt_video,
        'news': page_obj, 
        'videos': videos,
        'query': query
    })

# YANGI QO'SHILADIGAN FUNKSIYA (Dislike/Like uchun)
def news_reaction(request, pk, action):
    if request.method == "POST":
        news_item = get_object_or_404(News, pk=pk)
        if action == 'like':
            news_item.likes += 1
        elif action == 'dislike':
            news_item.dislikes += 1
        news_item.save()
        
        # Yangi sonlarni qaytaramiz
        return JsonResponse({
            'likes': news_item.likes,
            'dislikes': news_item.dislikes
        })
    



def news_detail(request, pk):
    item = get_object_or_404(News, pk=pk)
    yt_video = YouTubeVideo.objects.filter(is_active=True).last()
    item.views_count += 1
    item.save()
    context = {
        'item': item,
        'yt_video': yt_video,
        'full_url': request.build_absolute_uri() 
    }
    return render(request, 'core/news_detail.html', {'item': item})





def toggle_bookmark(request, news_id):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'login_required'}, status=401)

    if request.method == 'POST':
        news_item = get_object_or_404(News, id=news_id)
        bookmark, created = Bookmark.objects.get_or_create(user=request.user, news=news_item)
        
        if not created:
            bookmark.delete()
            return JsonResponse({'status': 'removed'})
        
        return JsonResponse({'status': 'added'})
    
    return JsonResponse({'status': 'error'}, status=400)




from django.contrib.auth.decorators import login_required

from django.core.paginator import Paginator
from django.db.models import Q
from student_management.models import Bookmark

@login_required
@login_required
def saved_news(request):
    # 1. Admin va oddiy user uchun filtr mantiqi
    if request.user.is_superuser:
        # Admin barcha foydalanuvchilar saqlagan hamma narsani ko'radi
        query = Bookmark.objects.all().select_related('news', 'user').order_by('-created_at')
    else:
        # Oddiy foydalanuvchi faqat o'zi saqlaganlarini ko'radi
        query = Bookmark.objects.filter(user=request.user).select_related('news').order_by('-created_at')

    # 2. Qidiruv mantiqi (Admin uchun user bo'yicha qidirish imkonini ham beradi)
    search_query = request.GET.get('q')
    if search_query:
        if request.user.is_superuser:
            query = query.filter(
                Q(news__title__icontains=search_query) | 
                Q(user__username__icontains=search_query) # Admin user nomi bo'yicha ham qidira oladi
            )
        else:
            query = query.filter(Q(news__title__icontains=search_query))

    # 3. Pagination
    paginator = Paginator(query, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'main/saved_news.html', {
        'bookmarks': page_obj,
        'total_count': query.count(),
        'search_query': search_query
    })



@login_required(login_url='core:login')
def add_news_frontend(request):
    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES)
        if form.is_valid():
            news = form.save(commit=False)
            news.author = request.user
            news.save()
            messages.success(request, "Yangilik qo'shildi!")
            return redirect('core:news_list')
    else:
        form = NewsForm()
    return render(request, 'core/add_news.html', {'form': form})

@login_required(login_url='core:login')
def like_news(request, pk):
    news = get_object_or_404(News, pk=pk)
    if request.user in news.likes.all():
        news.likes.remove(request.user)
    else:
        news.likes.add(request.user)
    return redirect(request.META.get('HTTP_REFERER', 'core:news_list'))

@login_required(login_url='core:login')
def add_comment(request, pk):
    news = get_object_or_404(News, pk=pk)
    if request.method == 'POST':
        text = request.POST.get('comment_text')
        if text:
            Comment.objects.create(news=news, user=request.user, text=text)
    return redirect('core:news_detail', pk=pk)

@login_required(login_url='core:login')
def add_video(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        v_url = request.POST.get('video_url')
        v_file = request.FILES.get('video_file')
        
        # 1. Ma'lumotlar kelganini tekshiramiz (Validation)
        if not title or not description:
            messages.warning(request, "Sarlavha va tavsif to'ldirilishi shart!")
            return render(request, 'core/video_upload.html')

        try:
            # 2. Ma'lumotni saqlash
            VideoNews.objects.create(
                title=title,
                description=description,
                video_url=v_url,
                video_file=v_file,
                author=request.user
            )
            messages.success(request, "Video muvaffaqiyatli yuklandi!")
            return redirect('core:dashboard')
            
        except Exception as e:
            # Baza bilan bog'liq kutilmagan xatolik bo'lsa
            messages.error(request, f"Xatolik yuz berdi: {e}")
            
    return render(request, 'core/video_upload.html')


def contact(request):
    if request.method == "POST":
        f_name = request.POST.get('full_name')
        email = request.POST.get('email')
        subj = request.POST.get('subject')
        msg = request.POST.get('message')

        if f_name and email and subj and msg:
            ContactMessage.objects.create(
                full_name=f_name, email=email, subject=subj, message=msg
            )
            messages.success(request, "Xabaringiz yuborildi!")
            return redirect('/') 
    return render(request, 'core/contact.html')

@login_required(login_url='core:login')
def all_messages(request):
    query = request.GET.get('q')
    filter_type = request.GET.get('filter')
    
    # Faqat o'chirilmagan xabarlar
    messages_list = ContactMessage.objects.filter(is_deleted=False).order_by('-created_at')

    if query:
        messages_list = messages_list.filter(
            Q(full_name__icontains=query) | Q(subject__icontains=query) | Q(message__icontains=query)
        )
    
    if filter_type == 'unread':
        messages_list = messages_list.filter(is_read=False)
    elif filter_type == 'starred':
        messages_list = messages_list.filter(is_starred=True)

    paginator = Paginator(messages_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'main/messages.html', {
        'messages': page_obj, 
        'query': query,
        'filter_type': filter_type
    })

@login_required(login_url='core:login')
def read_message(request, pk):
    msg = get_object_or_404(ContactMessage, pk=pk)
    msg.is_read = True
    msg.save()
    return render(request, 'main/message_detail.html', {'msg': msg})

@login_required(login_url='core:login')
def toggle_star_message(request, pk):
    if request.method == "POST":
        msg = get_object_or_404(ContactMessage, pk=pk)
        msg.is_starred = not msg.is_starred
        msg.save()
        return JsonResponse({'status': 'success', 'is_starred': msg.is_starred})
    return JsonResponse({'status': 'error'}, status=400)

@login_required(login_url='core:login')
def reply_message(request, pk):
    if request.method == 'POST':
        msg = get_object_or_404(ContactMessage, pk=pk)
        reply_text = request.POST.get('reply_text')
        
        if reply_text:
            try:
                subject = f"Re: {msg.subject}"
                full_message = f"Assalomu alaykum {msg.full_name},\n\nSizning xabaringizga javob:\n{reply_text}\n\nHurmat bilan, Maktab Ma'muriyati."
                send_mail(subject, full_message, settings.EMAIL_HOST_USER, [msg.email])
                msg.is_read = True 
                msg.save()
                messages.success(request, f"Javob yuborildi!")
            except Exception as e:
                messages.error(request, f"Xatolik: {e}")
        return redirect('core:read_message', pk=pk)

# --- SAVAT (TRASH) MANTIQI ---

@login_required(login_url='core:login')
def delete_message(request, pk):
    """Savatga tashlash"""
    if request.user.is_superuser:
        msg = get_object_or_404(ContactMessage, pk=pk)
        msg.is_deleted = True
        msg.save()
        messages.warning(request, "Xabar savatga tashlandi.")
    return redirect('core:all_messages')

@login_required(login_url='core:login')
def trash_messages(request):
    """Savatdagi xabarlar ro'yxati"""
    trash_list = ContactMessage.objects.filter(is_deleted=True).order_by('-created_at')
    return render(request, 'main/trash.html', {'trash_messages': trash_list})

@login_required(login_url='core:login')
def restore_message(request, pk):
    """Savatdan qayta tiklash"""
    msg = get_object_or_404(ContactMessage, pk=pk)
    msg.is_deleted = False
    msg.save()
    messages.success(request, "Xabar qayta tiklandi.")
    return redirect('core:trash_messages')

@login_required(login_url='core:login')
def permanent_delete_message(request, pk):
    """Bazadan butunlay o'chirish"""
    if request.user.is_superuser:
        msg = get_object_or_404(ContactMessage, pk=pk)
        msg.delete()
        messages.error(request, "Xabar butunlay o'chirildi.")
    return redirect('core:trash_messages')


# 5. API VA BILDIRISHNOMALAR


def check_messages_api(request):
    if request.user.is_authenticated:
        # Faqat o'chirilmagan va o'qilmagan xabarlar soni
        count = ContactMessage.objects.filter(is_read=False, is_deleted=False).count()
        return JsonResponse({'unread_count': count})
    return JsonResponse({'unread_count': 0})


 # Formani import qilishni unutmang




@login_required(login_url='core:login')
@login_required
def timetable_view(request):
    query = request.GET.get('q', '') # Qidiruv matnini olish
    
    # Bazadan darslarni guruh nomi yoki xona bo'yicha filtrlash
    all_lessons = Timetable.objects.select_related('group').all()
    
    if query:
        all_lessons = all_lessons.filter(
            Q(group__name__icontains=query) | 
            Q(room__icontains=query)
        )
    
    all_lessons = all_lessons.order_by('day_of_week', 'start_time')
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    timetable_data = {}
    
    for day_index, day_name in enumerate(days, start=1):
        day_lessons = [l for l in all_lessons if l.day_of_week == day_index]
        if day_lessons:
            timetable_data[day_name] = day_lessons

    context = {
        'timetable_data': timetable_data,
        'today_name': datetime.now().strftime('%A'),
        'search_query': query, # Qidiruv matnini inputda saqlab qolish uchun
    }
    return render(request, 'main/timetable.html', context)


# ==========================================
# 6. MOLIYA VA TO'LOVLAR (CUSTOM INTERFACE)
# ==========================================

@login_required(login_url='core:login')
@login_required(login_url='core:login')
def payment_list(request):
    # Payment endi tepada import qilingani uchun NameError yo'qoladi
    payments = Payment.objects.all().select_related('student').order_by('-date')
    
    # Modalda talabalarni tanlash uchun Student ham kerak (u ham importda bo'lishi shart)
    students_list = Student.objects.filter(enrollment_status='active')

    paginator = Paginator(payments, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    return render(request, 'main/payment_list.html', {
        'payments': page_obj,
        'students_list': students_list
    })






from django.utils import timezone


@login_required(login_url='core:login')

def add_payment(request):
    """Yangi to'lov qabul qilish va saqlash"""
    
    if request.method == 'POST':
        student_id = request.POST.get('student')
        amount_raw = request.POST.get('amount')
        p_method = request.POST.get('payment_type')
        
        # 1. Ma'lumotlarni tekshirish
        if not student_id or not amount_raw:
            messages.error(request, "Iltimos, talaba va summani kiriting.")
            return redirect('core:add_payment')

        # 2. Sonli ma'lumotga o'tkazish
        try:
            amount = float(amount_raw)
        except (TypeError, ValueError):
            messages.error(request, "Summa noto'g'ri formatda kiritildi.")
            return redirect('core:add_payment')

        # 3. Talabani topish va to'lovni yaratish
        student = get_object_or_404(Student, id=student_id)
        
        try:
            # transaction ishlatish tavsiya etiladi, chunki signalda balans o'zgaryapti
            with transaction.atomic():
                Payment.objects.create(
                    student=student,
                    amount=amount,
                    payment_method=p_method
                )
            messages.success(request, f"{student.first_name} uchun {amount} so'm to'lov qabul qilindi.")
            return redirect('core:payment_list')
        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")
            return redirect('core:add_payment')

    # --- BU QISMI YETISHMAYOTGAN EDI (GET request) ---
    # Sahifaga birinchi marta kirganda talabalar ro'yxatini yuborish
    students = Student.objects.filter(enrollment_status='active')
    
    # Agar faol talabalar bo'lmasa, barchasini ko'rsatish
    if not students.exists():
        students = Student.objects.all()

    return render(request, 'main/add_payment.html', {'students': students})
    
    # ... qolgan kodlar

@login_required(login_url='core:login')
def payment_detail(request, pk):
    """To'lov cheki/ma'lumotlari"""
    payment = get_object_or_404(Payment, pk=pk)
    return render(request, 'finance/payment_detail.html', {'payment': payment})





# 1. Excelga Eksport qilish
def export_timetable_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="timetable.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Sarlavhalar
    columns = ['Day', 'Group', 'Start Time', 'End Time', 'Room']
    ws.append(columns)
    
    # Ma'lumotlar
    lessons = Timetable.objects.all().order_by('day_of_week', 'start_time')
    for lesson in lessons:
        ws.append([
            lesson.day_of_week,
            lesson.group.name,
            lesson.start_time.strftime('%H:%M'),
            lesson.end_time.strftime('%H:%M'),
            lesson.room
        ])
    
    wb.save(response)
    return response

# 2. Exceldan Import qilish
def import_timetable_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # 2-qatordan boshlab o'qiymiz (1-qator sarlavha)
        for row in ws.iter_rows(min_row=2, values_only=True):
            day, group_name, start, end, room = row
            group = Group.objects.filter(name=group_name).first()
            if group:
                Timetable.objects.create(
                    day_of_week=day,
                    group=group,
                    start_time=start,
                    end_time=end,
                    room=room
                )
        return HttpResponseRedirect(reverse('timetable'))
    return HttpResponse("Xatolik yuz berdi")

from django.shortcuts import render, redirect, get_object_or_404

@login_required
def add_task(request):
    if request.method == "POST":
        title = request.POST.get('title')
        if title:
            Task.objects.create(user=request.user, title=title)
    return redirect('core:dashboard')

@login_required
def delete_task(request, pk):
    task = get_object_or_404(Task, id=pk, user=request.user)
    task.delete()
    return redirect('core:dashboard')

@login_required
def toggle_task(request, pk):
    task = get_object_or_404(Task, id=pk, user=request.user)
    task.is_completed = not task.is_completed
    task.save()
    return redirect('core:dashboard')